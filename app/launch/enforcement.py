"""
Launch v1 Enforcement Router for GenoMAX²
==========================================
Pipeline enforcement ensuring only Launch v1 products (TIER 1 + TIER 2)
reach external systems (Design Export, Shopify).

HARD GUARDRAILS:
- All external outputs MUST filter: is_launch_v1 = TRUE AND supplier_status = 'ACTIVE'
- No fuzzy matching. No heuristic inclusion.
- Brain logic and research views remain unfiltered.

PAIRING POLICY (v3.28.0):
- REQUIRED_PAIR: Must have exactly 2 environments (MAXimo² + MAXima²)
- SINGLE_ENV_ALLOWED: Must have exactly 1 environment (gender-specific)

Endpoints:
- GET  /api/v1/qa/launch-v1/pairing - Validate environment pairing with policy
- GET  /api/v1/launch-v1/export/design - Excel export with LAUNCH_V1_SUMMARY tab
- GET  /api/v1/launch-v1/products - List Launch v1 products with base_handle
- GET  /api/v1/launch-v1/summary - Quick dashboard summary

v3.28.0 - Policy-aware pairing
"""

import os
import re
import json
import io
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional, Tuple

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
import psycopg2
from psycopg2.extras import RealDictCursor

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(tags=["launch-v1"])

DATABASE_URL = os.getenv("DATABASE_URL")


# ===== Constants =====

# Hard guardrail: Launch v1 scope query
LAUNCH_V1_SCOPE_FILTER = """
    is_launch_v1 = TRUE 
    AND supplier_status = 'ACTIVE'
"""

# Pairing policy values
PAIRING_POLICY_REQUIRED_PAIR = 'REQUIRED_PAIR'
PAIRING_POLICY_SINGLE_ENV = 'SINGLE_ENV_ALLOWED'

# Regex to strip environment suffix from shopify_handle
ENV_SUFFIX_PATTERN = re.compile(r'-maximo$|-maxima$', re.IGNORECASE)


# ===== Database Helpers =====

def get_db():
    """Get database connection."""
    try:
        return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    except Exception as e:
        print(f"Database connection error: {e}")
        return None


def derive_base_handle(shopify_handle: Optional[str]) -> Optional[str]:
    """
    Derive base_handle by stripping -maximo or -maxima suffix.
    
    Example:
        vitamin-d3-5000iu-maximo -> vitamin-d3-5000iu
        omega-3-epa-maxima -> omega-3-epa
    """
    if not shopify_handle:
        return None
    return ENV_SUFFIX_PATTERN.sub('', shopify_handle)


def now_iso() -> str:
    """Return current UTC timestamp in ISO format."""
    return datetime.now(timezone.utc).isoformat()


def check_column_exists(cursor, table: str, column: str) -> bool:
    """Check if a column exists in a table."""
    cursor.execute("""
        SELECT column_name FROM information_schema.columns 
        WHERE table_name = %s AND column_name = %s
    """, (table, column))
    return cursor.fetchone() is not None


# ===== Models =====

class PairingOffender(BaseModel):
    """A base product with incorrect environment count based on its policy."""
    base_handle: str
    pairing_policy: str
    env_count: int
    expected_count: int
    module_codes: List[str]
    shopify_handles: List[str]
    environments: List[str]
    reason: str


class PairingQAResult(BaseModel):
    """Result of Launch v1 pairing QA check."""
    overall_status: str = Field(description="PASS or FAIL")
    timestamp: str
    scope: Dict[str, Any]
    distribution: List[Dict[str, Any]]
    total_base_products: int
    total_modules: int
    maximo_count: int
    maxima_count: int
    offenders: List[Dict[str, Any]]
    invariants: Dict[str, bool]
    policy_aware: bool = True


class LaunchV1Summary(BaseModel):
    """Summary stats for Launch v1 export."""
    modules_count: int
    base_products_count: int
    maximo_rows: int
    maxima_rows: int
    tier_1_count: int
    tier_2_count: int
    pairing_status: str
    invariant_modules_equals_sum: bool
    invariant_pairs_complete: bool


# ===== Core Functions =====

def fetch_launch_v1_products() -> List[Dict[str, Any]]:
    """
    Fetch all Launch v1 products with HARD GUARDRAILS.
    
    Returns only:
    - is_launch_v1 = TRUE
    - supplier_status = 'ACTIVE'
    """
    conn = get_db()
    if not conn:
        raise HTTPException(status_code=500, detail="Database connection failed")
    
    try:
        cur = conn.cursor()
        
        # Check if pairing_policy column exists
        has_pairing_policy = check_column_exists(cur, 'os_modules_v3_1', 'pairing_policy')
        has_base_handle_col = check_column_exists(cur, 'os_modules_v3_1', 'base_handle')
        
        # Build dynamic column list
        extra_cols = ""
        if has_pairing_policy:
            extra_cols += ", pairing_policy"
        if has_base_handle_col:
            extra_cols += ", base_handle AS db_base_handle"
        
        cur.execute(f"""
            SELECT 
                module_code,
                product_name,
                shopify_handle,
                os_environment,
                os_layer,
                tier,
                biological_domain,
                net_quantity,
                fda_disclaimer,
                shopify_body,
                front_label_text,
                back_label_text,
                suggested_use_full,
                safety_notes,
                contraindications,
                dosing_protocol,
                supplier_page_url,
                supliful_handle,
                supplier_status,
                is_launch_v1,
                created_at,
                updated_at
                {extra_cols}
            FROM os_modules_v3_1
            WHERE {LAUNCH_V1_SCOPE_FILTER}
            ORDER BY tier, os_environment, module_code
        """)
        
        rows = cur.fetchall()
        cur.close()
        conn.close()
        
        # Add derived base_handle if not from DB
        products = []
        for row in rows:
            product = dict(row)
            # Use DB base_handle if available, otherwise derive
            if has_base_handle_col and product.get('db_base_handle'):
                product['base_handle'] = product.pop('db_base_handle')
            else:
                product['base_handle'] = derive_base_handle(product.get('shopify_handle'))
                if 'db_base_handle' in product:
                    del product['db_base_handle']
            
            # Default pairing_policy if not in DB
            if not has_pairing_policy:
                product['pairing_policy'] = PAIRING_POLICY_REQUIRED_PAIR
            
            products.append(product)
        
        return products
        
    except Exception as e:
        try:
            conn.close()
        except:
            pass
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


def compute_policy_aware_pairing_analysis(products: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Analyze environment pairing for Launch v1 products using explicit policy.
    
    Policy rules:
    - REQUIRED_PAIR: Must have exactly 2 environments (MAXimo² + MAXima²)
    - SINGLE_ENV_ALLOWED: Must have exactly 1 environment
    
    Also detects duplicates within same (base_handle, os_environment).
    """
    # Group by base_handle
    base_groups: Dict[str, List[Dict[str, Any]]] = {}
    
    for product in products:
        base_handle = product.get('base_handle')
        if not base_handle:
            continue
        
        if base_handle not in base_groups:
            base_groups[base_handle] = []
        base_groups[base_handle].append(product)
    
    # Analyze each base_handle group
    offenders: List[PairingOffender] = []
    distribution: Dict[str, Dict[int, int]] = {
        PAIRING_POLICY_REQUIRED_PAIR: {},
        PAIRING_POLICY_SINGLE_ENV: {},
    }
    
    for base_handle, group in base_groups.items():
        # Get policy (should be same for all in group, use first)
        policy = group[0].get('pairing_policy', PAIRING_POLICY_REQUIRED_PAIR)
        
        # Count distinct environments
        envs = list(set(p.get('os_environment') for p in group))
        env_count = len(envs)
        
        # Track distribution
        if policy not in distribution:
            distribution[policy] = {}
        distribution[policy][env_count] = distribution[policy].get(env_count, 0) + 1
        
        # Determine expected count based on policy
        if policy == PAIRING_POLICY_REQUIRED_PAIR:
            expected = 2
        elif policy == PAIRING_POLICY_SINGLE_ENV:
            expected = 1
        else:
            expected = 2  # Default to pair
        
        # Check for policy violations
        is_violation = False
        reason = ""
        
        if policy == PAIRING_POLICY_REQUIRED_PAIR and env_count != 2:
            is_violation = True
            if env_count < 2:
                reason = f"REQUIRED_PAIR has only {env_count} environment(s), needs both MAXimo² and MAXima²"
            else:
                reason = f"REQUIRED_PAIR has {env_count} environments, should have exactly 2"
        
        elif policy == PAIRING_POLICY_SINGLE_ENV and env_count != 1:
            is_violation = True
            reason = f"SINGLE_ENV_ALLOWED has {env_count} environments, should have exactly 1"
        
        # Check for duplicates within same environment (regardless of policy)
        env_module_counts: Dict[str, int] = {}
        for p in group:
            env = p.get('os_environment', '')
            env_module_counts[env] = env_module_counts.get(env, 0) + 1
        
        for env, count in env_module_counts.items():
            if count > 1:
                is_violation = True
                reason = f"Duplicate modules ({count}) in same environment ({env})"
                break
        
        if is_violation:
            offenders.append(PairingOffender(
                base_handle=base_handle,
                pairing_policy=policy,
                env_count=env_count,
                expected_count=expected,
                module_codes=[p.get('module_code', '') for p in group],
                shopify_handles=[p.get('shopify_handle', '') for p in group],
                environments=envs,
                reason=reason
            ))
    
    # Count by environment
    maximo_count = sum(1 for p in products if p.get('os_environment') == 'MAXimo²')
    maxima_count = sum(1 for p in products if p.get('os_environment') == 'MAXima²')
    
    return {
        'base_groups': base_groups,
        'distribution': distribution,
        'offenders': offenders,
        'maximo_count': maximo_count,
        'maxima_count': maxima_count,
        'total_base_products': len(base_groups),
        'total_modules': len(products),
    }


# ===== Endpoints =====

@router.get("/api/v1/qa/launch-v1/pairing", response_model=PairingQAResult)
def qa_launch_v1_pairing():
    """
    QA Check: Validate Launch v1 has correct environment pairing per policy.
    
    POLICY RULES:
    - REQUIRED_PAIR: Must have exactly 2 environments (MAXimo² + MAXima²)
    - SINGLE_ENV_ALLOWED: Must have exactly 1 environment
    
    PASS conditions:
    - All REQUIRED_PAIR base_handles have exactly 2 environments
    - All SINGLE_ENV_ALLOWED base_handles have exactly 1 environment
    - No duplicate modules within same (base_handle, os_environment)
    
    FAIL conditions:
    - Any policy violation
    - Any duplicate within same environment
    
    Returns distribution by policy and list of offenders if any.
    """
    # Fetch products with hard guardrails
    products = fetch_launch_v1_products()
    
    # Compute policy-aware analysis
    analysis = compute_policy_aware_pairing_analysis(products)
    
    # Build distribution list
    dist_list = []
    for policy, env_counts in analysis['distribution'].items():
        for env_count, base_products in sorted(env_counts.items()):
            dist_list.append({
                "pairing_policy": policy,
                "env_count": env_count,
                "base_products": base_products
            })
    
    # Determine pass/fail
    offenders = analysis['offenders']
    overall_status = "PASS" if len(offenders) == 0 else "FAIL"
    
    # Invariant checks
    invariants = {
        "modules_equals_maximo_plus_maxima": (
            analysis['total_modules'] == analysis['maximo_count'] + analysis['maxima_count']
        ),
        "no_policy_violations": len(offenders) == 0,
    }
    
    return PairingQAResult(
        overall_status=overall_status,
        timestamp=now_iso(),
        scope={
            "is_launch_v1": True,
            "supplier_status": "ACTIVE"
        },
        distribution=dist_list,
        total_base_products=analysis['total_base_products'],
        total_modules=analysis['total_modules'],
        maximo_count=analysis['maximo_count'],
        maxima_count=analysis['maxima_count'],
        offenders=[o.model_dump() for o in offenders],
        invariants=invariants,
        policy_aware=True
    )


@router.get("/api/v1/launch-v1/products")
def list_launch_v1_products(
    environment: Optional[str] = Query(None, description="Filter by os_environment"),
    tier: Optional[str] = Query(None, description="Filter by tier (TIER 1, TIER 2)"),
    policy: Optional[str] = Query(None, description="Filter by pairing_policy"),
    limit: int = Query(500, ge=1, le=1000),
):
    """
    List all Launch v1 products with base_handle and pairing_policy.
    
    HARD GUARDRAILS enforced:
    - is_launch_v1 = TRUE
    - supplier_status = 'ACTIVE'
    """
    products = fetch_launch_v1_products()
    
    # Apply optional filters
    if environment:
        products = [p for p in products if p.get('os_environment') == environment]
    
    if tier:
        products = [p for p in products if p.get('tier') == tier]
    
    if policy:
        products = [p for p in products if p.get('pairing_policy') == policy]
    
    # Apply limit
    products = products[:limit]
    
    # Compute summary
    tier_counts = {}
    env_counts = {}
    policy_counts = {}
    for p in products:
        t = p.get('tier', 'UNKNOWN')
        e = p.get('os_environment', 'UNKNOWN')
        pol = p.get('pairing_policy', 'UNKNOWN')
        tier_counts[t] = tier_counts.get(t, 0) + 1
        env_counts[e] = env_counts.get(e, 0) + 1
        policy_counts[pol] = policy_counts.get(pol, 0) + 1
    
    return {
        "count": len(products),
        "timestamp": now_iso(),
        "scope": {
            "is_launch_v1": True,
            "supplier_status": "ACTIVE",
            "environment_filter": environment,
            "tier_filter": tier,
            "policy_filter": policy,
        },
        "tier_distribution": tier_counts,
        "environment_distribution": env_counts,
        "policy_distribution": policy_counts,
        "products": products,
    }


@router.get("/api/v1/launch-v1/export/design")
def export_design_excel(
    format: str = Query("xlsx", description="Export format (xlsx only for now)"),
):
    """
    Export Launch v1 products to Excel for Design team.
    
    HARD GUARDRAILS enforced:
    - is_launch_v1 = TRUE
    - supplier_status = 'ACTIVE'
    
    Includes:
    - LAUNCH_V1_SUMMARY tab with counts, policy distribution, and invariants
    - READY_FOR_DESIGN tab with all Launch v1 products
    - base_handle and pairing_policy columns
    
    Prevents "332 products" confusion by clearly showing:
    - modules_count (total rows)
    - base_products_count (unique base_handle)
    - breakdown by environment and policy
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=500, 
            detail="openpyxl not installed. Install with: pip install openpyxl"
        )
    
    # Fetch products with hard guardrails
    products = fetch_launch_v1_products()
    
    # Compute policy-aware analysis
    analysis = compute_policy_aware_pairing_analysis(products)
    
    # Compute tier counts
    tier_1_count = sum(1 for p in products if p.get('tier') == 'TIER 1')
    tier_2_count = sum(1 for p in products if p.get('tier') == 'TIER 2')
    
    # Policy counts
    required_pair_count = sum(1 for p in products if p.get('pairing_policy') == PAIRING_POLICY_REQUIRED_PAIR)
    single_env_count = sum(1 for p in products if p.get('pairing_policy') == PAIRING_POLICY_SINGLE_ENV)
    
    # Pairing status
    offenders = analysis['offenders']
    pairing_status = "PASS" if len(offenders) == 0 else "FAIL"
    
    # Invariant checks
    modules_equals_sum = analysis['total_modules'] == analysis['maximo_count'] + analysis['maxima_count']
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # ===== LAUNCH_V1_SUMMARY Tab =====
    ws_summary = wb.active
    ws_summary.title = "LAUNCH_V1_SUMMARY"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Summary data
    summary_data = [
        ("Launch v1 Export Summary (Policy-Aware)", ""),
        ("Generated At", now_iso()),
        ("", ""),
        ("COUNTS", ""),
        ("Total Modules (rows)", analysis['total_modules']),
        ("Unique Base Products", analysis['total_base_products']),
        ("", ""),
        ("ENVIRONMENT BREAKDOWN", ""),
        ("MAXimo² Rows", analysis['maximo_count']),
        ("MAXima² Rows", analysis['maxima_count']),
        ("", ""),
        ("TIER BREAKDOWN", ""),
        ("TIER 1 Count", tier_1_count),
        ("TIER 2 Count", tier_2_count),
        ("", ""),
        ("PAIRING POLICY BREAKDOWN", ""),
        ("REQUIRED_PAIR Modules", required_pair_count),
        ("SINGLE_ENV_ALLOWED Modules", single_env_count),
        ("", ""),
        ("INVARIANT CHECKS", ""),
        ("modules_count = maximo + maxima", "PASS" if modules_equals_sum else "FAIL"),
        ("Pairing Policy Compliance", pairing_status),
        ("Offender Count", len(offenders)),
        ("", ""),
        ("SCOPE FILTERS", ""),
        ("is_launch_v1", "TRUE"),
        ("supplier_status", "ACTIVE"),
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, start=1):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Style headers
        if label in ["Launch v1 Export Summary (Policy-Aware)", "COUNTS", "ENVIRONMENT BREAKDOWN", 
                     "TIER BREAKDOWN", "PAIRING POLICY BREAKDOWN", "INVARIANT CHECKS", "SCOPE FILTERS"]:
            ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # Style pass/fail
        if value == "PASS":
            ws_summary.cell(row=row_idx, column=2).fill = pass_fill
        elif value == "FAIL":
            ws_summary.cell(row=row_idx, column=2).fill = fail_fill
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 30
    
    # ===== READY_FOR_DESIGN Tab =====
    ws_design = wb.create_sheet("READY_FOR_DESIGN")
    
    # Headers
    headers = [
        "module_code",
        "product_name",
        "base_handle",
        "shopify_handle",
        "os_environment",
        "pairing_policy",
        "tier",
        "os_layer",
        "biological_domain",
        "net_quantity",
        "fda_disclaimer",
        "front_label_text",
        "back_label_text",
        "suggested_use_full",
        "safety_notes",
        "supplier_status",
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_design.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row_idx, product in enumerate(products, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = product.get(header, "")
            # Truncate long text
            if isinstance(value, str) and len(value) > 32000:
                value = value[:32000] + "..."
            ws_design.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        ws_design.column_dimensions[get_column_letter(col_idx)].width = min(40, max(12, len(header) + 2))
    
    # ===== PAIRING_OFFENDERS Tab (if any) =====
    if offenders:
        ws_offenders = wb.create_sheet("PAIRING_OFFENDERS")
        
        offender_headers = ["base_handle", "pairing_policy", "env_count", "expected_count", 
                           "module_codes", "environments", "reason"]
        for col_idx, header in enumerate(offender_headers, start=1):
            cell = ws_offenders.cell(row=1, column=col_idx, value=header)
            cell.fill = fail_fill
            cell.font = Font(bold=True)
        
        for row_idx, offender in enumerate(offenders, start=2):
            ws_offenders.cell(row=row_idx, column=1, value=offender.base_handle)
            ws_offenders.cell(row=row_idx, column=2, value=offender.pairing_policy)
            ws_offenders.cell(row=row_idx, column=3, value=offender.env_count)
            ws_offenders.cell(row=row_idx, column=4, value=offender.expected_count)
            ws_offenders.cell(row=row_idx, column=5, value=", ".join(offender.module_codes))
            ws_offenders.cell(row=row_idx, column=6, value=", ".join(offender.environments))
            ws_offenders.cell(row=row_idx, column=7, value=offender.reason)
        
        # Adjust widths
        ws_offenders.column_dimensions['A'].width = 35
        ws_offenders.column_dimensions['B'].width = 20
        ws_offenders.column_dimensions['E'].width = 50
        ws_offenders.column_dimensions['G'].width = 60
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"genomax2_launch_v1_design_export_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/api/v1/launch-v1/summary")
def get_launch_v1_summary():
    """
    Quick summary of Launch v1 scope without full product list.
    
    Useful for dashboard and status checks.
    """
    products = fetch_launch_v1_products()
    analysis = compute_policy_aware_pairing_analysis(products)
    
    tier_1_count = sum(1 for p in products if p.get('tier') == 'TIER 1')
    tier_2_count = sum(1 for p in products if p.get('tier') == 'TIER 2')
    
    offenders = analysis['offenders']
    pairing_status = "PASS" if len(offenders) == 0 else "FAIL"
    
    modules_equals_sum = analysis['total_modules'] == analysis['maximo_count'] + analysis['maxima_count']
    pairs_complete = pairing_status == "PASS"
    
    return LaunchV1Summary(
        modules_count=analysis['total_modules'],
        base_products_count=analysis['total_base_products'],
        maximo_rows=analysis['maximo_count'],
        maxima_rows=analysis['maxima_count'],
        tier_1_count=tier_1_count,
        tier_2_count=tier_2_count,
        pairing_status=pairing_status,
        invariant_modules_equals_sum=modules_equals_sum,
        invariant_pairs_complete=pairs_complete
    )
